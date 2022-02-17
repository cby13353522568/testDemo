# -*- coding: UTF-8 -*-
# @Time: 2021/1/25 21:47
# @File: openpyxl读写excel.py

from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook


def readExcel(path):
    file = load_workbook(filename=path)
    sheets = file.sheetnames
    dict = {}
    for sheetName in sheets:
        sheet = file[sheetName]
        sheetInfo = []
        for rowNum in range(1, sheet.max_row + 1):
            row = []
            for colNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=rowNum, column=colNum).value
                row.append(value)
            sheetInfo.append(row)
        dict[sheetName] = sheetInfo
    return dict


def writeExcel():
    workBook = Workbook()
    sheet = workBook.active
    # sheet['A1'] = 'new'
    sheet.cell(2, 3).value = "ok"  # 第二行第三列

    workBook.save('openpyxl写入excel.xlsx')

# openpyxl只能操作xlsx文件，当我们遇到xls文件的时候就需要进行转化
# import os
# import win32com.client as win32
# filename = r'C:\Users\XH\Desktop\1.xls'
# Excelapp = win32.gencache.EnsureDispatch('Excel.Application')
# workbook = Excelapp.Workbooks.Open(filename)
# # 转xlsx时: FileFormat=51,
# # 转xls时:  FileFormat=56,
# workbook.SaveAs(filename.replace('xls', 'xlsx'), FileFormat=51)
# workbook.Close()
# Excelapp.Application.Quit()
# # 删除源文件
# # os.remove(filename)
# # 如果想将xlsx的文件转换为xls的话，则可以使用以下的代码：
# # workbook.SaveAs(filename.replace('xlsx', 'xls'), FileFormat=56)


if __name__ == '__main__':
    path = r"E:\cui\C_TEST\test_cases\办公文件读取\1.xlsx"
    excel = readExcel(path)
    print(excel)
    # writeExcel()
