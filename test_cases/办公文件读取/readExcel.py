# -*- coding: UTF-8 -*-
# @Time: 2021/1/25 21:47
# @File: readExcel.py

from openpyxl.reader.excel import load_workbook

def readExcel(path):
    file = load_workbook(filename=path)
    sheets = file.sheetnames
    dict={}
    for sheetName in sheets:
        sheet=file[sheetName]
        sheetInfo=[]
        for rowNum in range(1, sheet.max_column):
            row = []
            for colNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=rowNum, column=colNum).value
                row.append(value)
            sheetInfo.append(row)
        dict[sheetName] = sheetInfo
    return dict


path=r"E:\cui\C_TEST\test_cases\办公文件读取\1.xlsx"
excel = readExcel(path)
print(excel)